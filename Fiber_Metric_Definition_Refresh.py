#### FIBER METRIC DEFINITION REFRESH SCRIPT ####
### CREATED BY CARSON GARLAND (cg119y) ###

### IMPORTING REQUIRED PACKAGES ###
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from re import sub
import win32com.client
from pywintypes import com_error
from glob import glob
import os
from datetime import datetime
from pick import pick

### DICTS AND LISTS ###
## DICTS FOR SELECTING OPTIONS ##
org_options_list = [
    'ALL', 
    'Consumer', 
    'Business', 
    'NEO'
    ]

full_metric_options_list = [
    'ALL',
    'Consolidated',
    'Fiber Gross Adds',
    'Fiber Churn',
    'Fiber Net Adds',
    'Fiber Gross Adds Speed Mix',
    'Fiber Penetration',
    'Fiber Customer Base',
    'PPV Penetration of Fiber HH Base',
    'Fiber Penetration of Eligible PPV',
    'NPS',
    'Reliability Satisfaction Score',
    'Cust. Service Satisfaction Score',
    'Price Value Satisfaction Score',
    'Billing Satisfaction Score',
    'Fiber Revenue',
    'Fiber ARPU',
    'Net Migrations',
    'Fiber Overbuild 2022 ELUs',
    'Net Adds from 2022 Overbuild',
    'Fiber Gross Adds',
    'IPBB Gross Adds Speed Mix',
    'Fiber Churn',
    'Fiber Net Adds',
    'Fiber Penetration',
    'Total Fiber Ports EOP',
    'Lit Buildings',
    'Fiber Lit BCLs',
    'In-Franchise Fiber Coverage',
    'Fiber Connectivity Revenue',
    'Fiber ARPU',
    'Fiber Yield of Service',
    'CLs Passed Greenfield',
    'CLs Passed Overbuild',
    'CPCL - FTTP (Greenfield)',
    'CPCL - IFP (Overbuild)'
]

consumer_metric_options_list = [
    'ALL',
    'Consolidated',
    'Fiber Gross Adds',
    'Fiber Churn',
    'Fiber Net Adds',
    'Fiber Gross Adds Speed Mix',
    'Fiber Penetration',
    'Fiber Customer Base',
    'PPV Penetration of Fiber HH Base',
    'Fiber Penetration of Eligible PPV',
    'NPS',
    'Reliability Satisfaction Score',
    'Cust. Service Satisfaction Score',
    'Price Value Satisfaction Score',
    'Billing Satisfaction Score',
    'Fiber Revenue',
    'Fiber ARPU',
    'Net Migrations',
    'Fiber Overbuild 2022 ELUs',
    'Net Adds from 2022 Overbuild'
]

business_metric_options_list = [
    'ALL',
    'Consolidated',
    'Fiber Gross Adds',
    'IPBB Gross Adds Speed Mix',
    'Fiber Churn',
    'Fiber Net Adds',
    'Fiber Penetration',
    'Total Fiber Ports EOP',
    'Lit Buildings',
    'Fiber Lit BCLs',
    'In-Franchise Fiber Coverage',
    'Fiber Connectivity Revenue',
    'Fiber ARPU',
    'Fiber Yield of Service'
]

neo_metric_options_list = [
    'ALL',
    'Consolidated',
    'CLs Passed Greenfield',
    'CLs Passed Overbuild',
    'CPCL - FTTP (Greenfield)',
    'CPCL - IFP (Overbuild)'
]

## DICT TO CONVERT FROM EXCEL METRIC NAME TO PDF FILE NAME ##
# 'EXCEL METRIC NAME : PDF FILE NAME' #
metric_exceltopdf_dict = {
    'Connectivity Revenue':'Connectivity Revenue',
    'Consolidated':'Consolidated',
    'Fiber ARPU':'Fiber_ARPU',
    'Fiber Churn':'Fiber_Churn',
    'Fiber Customer BCLs':'Fiber_Customer BCLs',
    'Fiber Gross Adds':'Fiber_GrossAdds',
    'Fiber Net Adds':'Fiber_NetAdds',
    'Fiber Penetration':'Fiber_Penetration',
    'Fiber Yield of Service':'Fiber_Yield_of_Service',
    'Total Fiber Ports EOP':'Total_Fiber_Ports_EOP',
    'Lit Buildings':'Lit_Buildings',
    'Fiber Customer Base':'Fiber_Customer_Base',
    'Fiber Gross Adds Speed Mix':'Fiber_GrossAdds_SpeedMix',
    'Fiber Overbuild 2022 ELUs':'Fiber_Overbuild_2022_ELUs',
    'Fiber Penetration of Eligible PPV':'Fiber_Penetration_of_Eligible_PPV',
    'Fiber Revenue':'Fiber_Revenue',
    'Net Adds from 2022 Overbuild':'Net_Adds_from_2022_Overbuild',
    'Net Migrations':'Net_Migrations',
    'NPS':'NPS',
    'PPV Penetration of Fiber HH Base':'PPV_Penetration_of_Fiber_HHBase',
    'Reliability Satisfaction Score':'NPS-Reliability_Satisfaction_Score',
    'Cust. Service Satisfaction Score':'NPS-Customer_Service_Satisfaction_Score',
    'Price Value Satisfaction Score':'NPS-Price_Value_Satisfaction_Score',
    'Billing Satisfaction Score':'NPS-Billing_Satisfaction_Score',
    'IPBB Gross Adds Speed Mix':'IPBB_GrossAdds_SpeedMix',
    'Fiber Lit BCLs':'Fiber_Lit BCLs',
    'In-Franchise Fiber Coverage':'In-Franchise_Fiber_Coverage',
    'Fiber Connectivity Revenue':'Fiber_Connectivity_Revenue',
    'CLs Passed Greenfield':'CLs_Passed_Greenfield',
    'CLs Passed Overbuild':'CLs_Passed_Overbuild',
    'CPCL - FTTP (Greenfield)':'CPCL_Greenfield',
    'CPCL - IFP (Overbuild)':'CPCL_Overbuild'
    }

## LIST OF METRIC INFO TO INCLUDE ON EACH REPORT ##
metric_info_list = [
    'Metric Definition', 
    'Scope', 
    #'Current State Metric Breakdown', 
    'Current State Metric Source', 
    #'Current State Drilldown', 
    'Source Contact / Metric Owner / Business Owner', 
    #'Categories / Trends / Benchmarks',
    #'Cards',
    'Unit of Measure',
    'More Information on Metric'
    ]

### SAVE EXCEL FILE AS PDF ###
def saveaspdf(filename):
    
    WB_PATH = filename
    PATH_TO_PDF = WB_PATH.replace('xlsx', 'pdf')

    # OPEN EXCEL APP IN THE BACKGROUND #
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False

    try:

        # OPEN INPUT EXCEL FILE #
        wb = excel.Workbooks.Open(WB_PATH)

        sh_idx = 1
        ws_index_list = []

        # FORMATTING PAGE TITLE AND RELEASE DATE #
        for sh in wb.Sheets:
            myRange1 = sh.Range(sh.Cells(1,1),sh.Cells(1,2))
            myRange1.MergeCells = True
            myRange2 = sh.Range(sh.Cells(metric_num + 3,1),sh.Cells(metric_num + 3,2))
            myRange2.MergeCells = True
            ws_index_list.append(sh_idx)
            wb.WorkSheets(ws_index_list).Select()
            sh_idx+=1

        # SAVE AS PDF #
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)

    except com_error as e:
        print('failed.')
    finally:

        # CLOSE EXCEL AND REMOVE EXCEL FILE #
        wb.Close()
        os.remove(WB_PATH) #TO KEEP THE EXCEL FILES AS WELL, COMMENT OUT THIS LINE
        excel.ScreenUpdating = True
        excel.DisplayAlerts = True
        excel.EnableEvents = True

### APPEND METRIC INFO TO EXCEL SHEET ###
def metricstosheet(wso1, wsi2, metric_info_dict, metric, org, sh_org, in_row):
    
    global metric_num
    
    # APPEND CONSOLIDATED METRIC INFO TO OUTPUT SHEET#
    wso1.append([wsi2.cell(row=2, column=5).value]) #Page Title
    wso1.append([org + ' Fiber', sub("\*", '', metric)]) #Table Title

    # METRIC INFO START, IF REMOVED OR ADDED TO, UPDATE METRIC_NUM ABOVE #
    metric_num = 0
    for key in metric_info_dict:
        if key in metric_info_list:
            wso1.append([key, str(in_row[metric_info_dict.get(key)].value)])
            metric_num += 1
    # METRIC INFO END #

    wso1.append([wsi2.cell(row=4, column=5).value]) #Release Date
    
    # FORMATTING TABLE #
    for out_row in wso1.rows:
        for cell in out_row:
            cell.alignment = Alignment(vertical= 'top', wrapText=True)
    wso1.cell(row=1, column=1).style = 'Headline 2'
    wso1.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    wso1.cell(row=2, column=1).font = Font(size=12, color='FFFFFFFF')
    wso1.cell(row=2, column=2).font = Font(size=12, color='FFFFFFFF')
    wso1.cell(row=2, column=1).alignment = Alignment(vertical= 'center', wrapText=True)
    wso1.cell(row=2, column=2).alignment = Alignment(vertical= 'center', wrapText=True)
    wso1.column_dimensions['A'].width = 15.5
    wso1.column_dimensions['B'].width = 69
    
    if(sh_org):
        sh_org += metric

    # CREATE TABLE FOR FORMATTING #
    tab_ref = 'B' + str((metric_num + 2))
    tab = Table(displayName= sub('[^a-zA-Z0-9\.]', '', sh_org)[:30], ref="A2:"+tab_ref)
    # CHANGE TABLE STYLE HERE #
    style = TableStyleInfo(name='TableStyleMedium10', showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    wso1.add_table(tab)

### SAVE IDV AND CONSOLIDATED EXCEL FILES ###
def saveasexcel(input_filename):

    ## CREATING REPORT FOLDER ##
    path = os.path.abspath("Fiber_Metric_Definition_Refresh.exe")
    path = path.rsplit('\\', 1)[0]
    date = str(datetime.now())
    date = date.rsplit(' ', 1)[0]
    report_path = path + '\\Fiber Metric Definitions ' + date
    while True:
        try:
            os.mkdir(report_path)
        except OSError as error:
            print(error)
            input("Please delete existing 'Fiber Metric Defintions [todays date]' folder and press Enter to continue: ")
            continue
        else:
            break

    ## LOADING INPUT EXCEL FILE ##
    wbi1 = load_workbook(filename = input_filename)
    wsi1 = wbi1['LIST']
    wsi2 = wbi1['About']
    
    ## CREATE DICTS OF HEADER INFO TO INDEX FROM ##
    headers_dict = {}
    metric_info_dict = {}
    for cell in wsi1[1]:
        headers_dict[cell.value] = cell.col_idx - 1
    for info in metric_info_list:
        for cell in wsi1[1]:
            if(cell.value == info):
                metric_info_dict[info] = cell.col_idx - 1

    ## PROMPTING USER FOR WHICH REPORTS THEY WOULD LIKE ##
    metrics_selected = []
    orgs_selected = []

    title = 'Please choose which organizations you want to create metric definitions for (press SPACE to mark, ENTER to continue): '
    orgs_picked = pick(org_options_list, title, multiselect=True, min_selection_count=1)

    if(orgs_picked[0][0] == 'ALL'):
        orgs_selected = ['Consumer', 'Business', 'NEO']
        title = 'Please choose which metric definitions you want to create (press SPACE to mark, ENTER to continue): '
        selected = pick(full_metric_options_list, title, multiselect=True, min_selection_count=1)
        if(selected[0][0] == 'ALL'):
            for option in consumer_metric_options_list:
                    metrics_selected.append(('Consumer', option))
            for option in business_metric_options_list:
                    metrics_selected.append(('Business', option))
            for option in neo_metric_options_list:
                metrics_selected.append(('NEO', option))
        else:
            for option in selected:
                metrics_selected.append(('Consumer', option[0]))
                metrics_selected.append(('Business', option[0]))
                metrics_selected.append(('NEO', option[0]))
    else:
        for org_tuple in orgs_picked:
            orgs_selected.append(org_tuple[0])

        if 'Consumer' in orgs_selected:
            title = 'Please choose which Consumer metric definitions to create (press SPACE to mark, ENTER to continue): '
            selected = pick(consumer_metric_options_list, title, multiselect=True, min_selection_count=1)
            if(selected[0][0] == 'ALL'):
                for option in consumer_metric_options_list:
                    if(option != 'ALL'):
                        metrics_selected.append(('Consumer', option))
            else:
                for selected_option in selected:
                    metrics_selected.append(('Consumer', selected_option[0]))
                    
        if 'Business' in orgs_selected:
            title = 'Please choose which Business metric definitions create (press SPACE to mark, ENTER to continue): '
            selected = pick(business_metric_options_list, title, multiselect=True, min_selection_count=1)
            if(selected[0][0] == 'ALL'):
                for option in business_metric_options_list:
                    if(option != 'ALL'):
                        metrics_selected.append(('Business', option))
            else:
                for selected_option in selected:
                    metrics_selected.append(('Business', selected_option[0]))

        if 'NEO' in orgs_selected:
            title = 'Please choose which NEO metric definitions to create (press SPACE to mark, ENTER to continue): '
            selected = pick(neo_metric_options_list, title, multiselect=True, min_selection_count=1)
            if(selected[0][0] == 'ALL'):
                for option in neo_metric_options_list:
                    if(option != 'ALL'):
                        metrics_selected.append(('NEO', option))
            else:
                for selected_option in selected:
                    metrics_selected.append(('NEO', selected_option[0]))

    ## CREATES EXCEL FILES FOR SELECTED METRICS ##
    for org in orgs_selected:
        
        # CREATE OUTPUT WORKBOOK FOR CONSOLIDATED PDF #
        wbo1 = Workbook()

        # ITERATE THROUGH ROWS (METRICS) #
        for in_row in wsi1.rows:
            metric = in_row[headers_dict.get('Metric')].value
            sh_org = in_row[headers_dict.get('Organization')].value
            if(((org, metric) in metrics_selected or (org, 'Consolidated') in metrics_selected) and sh_org == org and metric_exceltopdf_dict.get(metric)):
                
                # FOR CONSOLIDATED PDF #
                # CREATE OUTPUT SHEET #
                wso1 = wbo1.create_sheet(title= sub('[^a-zA-Z0-9 \n\.]', ' ', metric)[:30])
                
                # APPEND CONSOLIDATED METRIC INFO TO OUTPUT SHEET #
                metricstosheet(wso1, wsi2, metric_info_dict, metric, org, sh_org, in_row)

                ## FOR IDV EXCEL FILES ##
                # CREATE OUTPUT WORKBOOK FOR IDV PDF #
                wbo2 = Workbook()

                # CREATE OUTPUT SHEET #
                wso2 = wbo2.create_sheet(title= sub('[^a-zA-Z0-9 \n\.]', ' ', metric)[:30])

                # APPEND IDV METRIC INFO TO OUTPUT SHEET #
                metricstosheet(wso2, wsi2, metric_info_dict, metric, org, sh_org, in_row)
                
                del wbo2['Sheet']

                # SAVE IDV EXCEL FILES AND PASS THEM TO SAVEASPDF() #
                if(metric_exceltopdf_dict.get(metric) == 'Lit_Buildings' and (org, metric) in metrics_selected):
                    print('SAVING : ' + org + ' ' + metric)
                    idv_excel_filename = report_path + '\\FiberDashboard-Definitions-' + org + '-' + 'Lit_Buildings.xlsx'
                    wbo2.save(idv_excel_filename)
                    saveaspdf(idv_excel_filename)                   
                    idv_excel_filename = report_path + '\\FiberDashboard-Definitions-' + org + '-' + 'Total_Lit_IF_Nationally.xlsx'
                    wbo2.save(idv_excel_filename)
                    saveaspdf(idv_excel_filename)
                elif((org, metric) in metrics_selected):
                    print('SAVING : ' + org + ' ' + metric)
                    idv_excel_filename = report_path + '\\FiberDashboard-Definitions-' + org + '-' + metric_exceltopdf_dict.get(metric) + '.xlsx'
                    wbo2.save(idv_excel_filename)
                    saveaspdf(idv_excel_filename)

        del wbo1['Sheet']

        # SAVE CONSOLIDATED EXCEL FILE AND PASS IT TO SAVEASPDF() #
        if((org, 'Consolidated') in metrics_selected):
            print('SAVING : Consolidated ' + org)
            excel_filename = report_path + '\\FiberDashboard-Definitions-' + org + '-Consolidated.xlsx'
            wbo1.save(excel_filename)
            saveaspdf(excel_filename)

### MAIN METHOD ###
def main():
    
    # GET THE INPUT EXCEL FILE AND PASS IT TO SAVEASEXCEL() #
    while True:
        try:
            input_filename = glob('Executive Fiber Metric Dashboard-Metric Definitions PDF Printout*.xlsx')[0]
        except IndexError as error:
            input("Please ensure the 'Executive Fiber Metric Dashboard-Metric Definitions PDF Printout*.xlsx' file is in the same directory as the executable, then press Enter to continue: ")
            continue
        else:
            break
    print('RUNNING...')
    saveasexcel(input_filename)
    print('COMPLETE!')

if __name__ == "__main__":
    main()